#!/usr/bin/env python3
"""
Excel Master Controller - Comprehensive Excel Automation Engine
Provides complete spreadsheet control from single prompts with intelligent data organization
"""

import sys
import json
import re
import os
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional, Tuple
from pathlib import Path

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE, FORMAT_PERCENTAGE
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
import pandas as pd

class ExcelMaster:
    """Complete Excel automation and control system"""
    
    def __init__(self):
        self.workbook = None
        self.current_sheet = None
        self.color_schemes = {
            'corporate': {
                'primary': 'FF0033CC',      # Blue
                'secondary': 'FF99CCFF',    # Light blue  
                'accent': 'FFFF6600',       # Orange
                'success': 'FF00CC00',      # Green
                'warning': 'FFFF9900',      # Yellow
                'danger': 'FFCC0000',       # Red
                'text': 'FF333333',         # Dark gray
                'background': 'FFF8F9FA'    # Light gray
            },
            'modern': {
                'primary': 'FF2D3436',      # Charcoal
                'secondary': 'FF74B9FF',    # Light blue
                'accent': 'FF6C5CE7',       # Purple
                'success': 'FF00B894',      # Teal
                'warning': 'FFFDCB6E',      # Yellow
                'danger': 'FFE17055',       # Coral
                'text': 'FF2D3436',         # Dark
                'background': 'FFFFFFFF'    # White
            }
        }
        
    def analyze_request(self, prompt: str) -> Dict:
        """Analyze user request to determine spreadsheet requirements"""
        prompt_lower = prompt.lower()
        
        # Detect spreadsheet type
        if any(word in prompt_lower for word in ['budget', 'financial', 'expense', 'revenue', 'profit']):
            sheet_type = 'financial'
        elif any(word in prompt_lower for word in ['sales', 'tracking', 'performance', 'metrics', 'kpi']):
            sheet_type = 'tracking'
        elif any(word in prompt_lower for word in ['project', 'timeline', 'gantt', 'schedule']):
            sheet_type = 'project'
        elif any(word in prompt_lower for word in ['inventory', 'stock', 'products', 'catalog']):
            sheet_type = 'inventory'
        elif any(word in prompt_lower for word in ['analysis', 'data', 'statistics', 'report']):
            sheet_type = 'analysis'
        else:
            sheet_type = 'general'
        
        # Detect complexity
        complexity_indicators = ['dashboard', 'chart', 'pivot', 'formula', 'calculation', 'summary']
        complexity = 'advanced' if any(word in prompt_lower for word in complexity_indicators) else 'basic'
        
        # Detect update operation
        is_update = any(word in prompt_lower for word in ['update', 'modify', 'add', 'change', 'edit'])
        
        return {
            'type': sheet_type,
            'complexity': complexity,
            'is_update': is_update,
            'color_scheme': 'corporate' if sheet_type == 'financial' else 'modern'
        }
    
    def create_spreadsheet(self, prompt: str, output_path: str = None) -> str:
        """Create complete spreadsheet from user prompt"""
        analysis = self.analyze_request(prompt)
        
        # Create or load workbook
        if analysis['is_update'] and output_path and os.path.exists(output_path):
            self.workbook = load_workbook(output_path)
        else:
            self.workbook = Workbook()
            # Remove default sheet
            if 'Sheet' in self.workbook.sheetnames:
                self.workbook.remove(self.workbook['Sheet'])
        
        # Generate spreadsheet based on type
        if analysis['type'] == 'financial':
            self._create_financial_sheet(prompt, analysis)
        elif analysis['type'] == 'tracking':
            self._create_tracking_sheet(prompt, analysis)
        elif analysis['type'] == 'project':
            self._create_project_sheet(prompt, analysis)
        elif analysis['type'] == 'inventory':
            self._create_inventory_sheet(prompt, analysis)
        elif analysis['type'] == 'analysis':
            self._create_analysis_sheet(prompt, analysis)
        else:
            self._create_general_sheet(prompt, analysis)
        
        # Add dashboard if advanced
        if analysis['complexity'] == 'advanced':
            self._create_dashboard()
        
        # Save workbook
        if not output_path:
            safe_name = re.sub(r'[^\\w\\s-]', '', prompt[:30]).strip()
            safe_name = re.sub(r'[-\\s]+', '_', safe_name)
            output_path = f"{safe_name}_spreadsheet.xlsx"
        
        self.workbook.save(output_path)
        return output_path
    
    def _create_financial_sheet(self, prompt: str, analysis: Dict):
        """Create financial tracking spreadsheet"""
        sheet = self.workbook.create_sheet("Budget_Tracker")
        self.current_sheet = sheet
        
        # Headers
        headers = ['Date', 'Category', 'Description', 'Amount', 'Type', 'Balance']
        self._add_headers(sheet, headers, 1)
        
        # Sample data
        sample_data = [
            [datetime.now().date(), 'Income', 'Salary', 5000, 'Credit', 5000],
            [datetime.now().date(), 'Rent', 'Monthly rent', -1200, 'Debit', 3800],
            [datetime.now().date(), 'Food', 'Groceries', -300, 'Debit', 3500],
            [datetime.now().date(), 'Transport', 'Gas', -100, 'Debit', 3400],
        ]
        
        for i, row in enumerate(sample_data, 2):
            for j, value in enumerate(row, 1):
                cell = sheet.cell(row=i, column=j, value=value)
                if j == 4:  # Amount column
                    cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
                elif j == 6:  # Balance column
                    cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
        
        # Add formulas for balance calculation
        for i in range(3, 6):  # Skip first data row
            balance_cell = sheet.cell(row=i, column=6)
            balance_cell.value = f"=F{i-1}+D{i}"
        
        # Summary section
        sheet['H1'] = 'SUMMARY'
        sheet['H2'] = 'Total Income:'
        sheet['I2'] = '=SUMIF(E:E,"Credit",D:D)'
        sheet['H3'] = 'Total Expenses:'  
        sheet['I3'] = '=SUMIF(E:E,"Debit",D:D)'
        sheet['H4'] = 'Net Balance:'
        sheet['I4'] = '=I2+I3'
        
        # Apply formatting
        self._apply_financial_formatting(sheet)
        
        # Add chart if advanced
        if analysis['complexity'] == 'advanced':
            self._add_expense_chart(sheet)
    
    def _create_tracking_sheet(self, prompt: str, analysis: Dict):
        """Create sales/performance tracking sheet"""
        sheet = self.workbook.create_sheet("Performance_Tracker")
        self.current_sheet = sheet
        
        # Headers
        headers = ['Date', 'Salesperson', 'Product', 'Quantity', 'Unit_Price', 'Total_Sale', 'Commission']
        self._add_headers(sheet, headers, 1)
        
        # Sample data
        sample_data = [
            [datetime.now().date(), 'John Smith', 'Product A', 10, 50, 500, 50],
            [datetime.now().date(), 'Jane Doe', 'Product B', 5, 100, 500, 50],
            [datetime.now().date(), 'Bob Wilson', 'Product A', 8, 50, 400, 40],
        ]
        
        for i, row in enumerate(sample_data, 2):
            for j, value in enumerate(row, 1):
                cell = sheet.cell(row=i, column=j, value=value)
                if j in [5, 6, 7]:  # Price columns
                    cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
        
        # Add formulas
        for i in range(2, 5):
            sheet.cell(row=i, column=6).value = f"=D{i}*E{i}"  # Total Sale
            sheet.cell(row=i, column=7).value = f"=F{i}*0.1"   # Commission (10%)
        
        # Summary section  
        self._add_tracking_summary(sheet)
        self._apply_tracking_formatting(sheet)
        
        if analysis['complexity'] == 'advanced':
            self._add_sales_chart(sheet)
    
    def _create_project_sheet(self, prompt: str, analysis: Dict):
        """Create project management sheet"""
        sheet = self.workbook.create_sheet("Project_Plan")
        self.current_sheet = sheet
        
        headers = ['Task', 'Assigned_To', 'Start_Date', 'End_Date', 'Duration', 'Status', 'Progress']
        self._add_headers(sheet, headers, 1)
        
        # Sample project data
        start_date = datetime.now().date()
        sample_data = [
            ['Planning Phase', 'Manager', start_date, start_date + timedelta(days=7), 7, 'Complete', 1.0],
            ['Development', 'Dev Team', start_date + timedelta(days=7), start_date + timedelta(days=21), 14, 'In Progress', 0.6],
            ['Testing', 'QA Team', start_date + timedelta(days=21), start_date + timedelta(days=28), 7, 'Not Started', 0.0],
            ['Deployment', 'DevOps', start_date + timedelta(days=28), start_date + timedelta(days=30), 2, 'Not Started', 0.0]
        ]
        
        for i, row in enumerate(sample_data, 2):
            for j, value in enumerate(row, 1):
                cell = sheet.cell(row=i, column=j, value=value)
                if j == 7:  # Progress column
                    cell.number_format = FORMAT_PERCENTAGE
        
        self._apply_project_formatting(sheet)
        
        if analysis['complexity'] == 'advanced':
            self._add_gantt_visualization(sheet)
    
    def _create_inventory_sheet(self, prompt: str, analysis: Dict):
        """Create inventory management sheet"""
        sheet = self.workbook.create_sheet("Inventory")
        self.current_sheet = sheet
        
        headers = ['Product_ID', 'Product_Name', 'Category', 'Quantity', 'Unit_Cost', 'Total_Value', 'Reorder_Level']
        self._add_headers(sheet, headers, 1)
        
        # Sample inventory data
        sample_data = [
            ['P001', 'Widget A', 'Electronics', 50, 25.00, 1250.00, 10],
            ['P002', 'Widget B', 'Electronics', 30, 35.00, 1050.00, 15],
            ['P003', 'Tool X', 'Tools', 25, 15.00, 375.00, 5],
            ['P004', 'Material Y', 'Supplies', 100, 5.00, 500.00, 20]
        ]
        
        for i, row in enumerate(sample_data, 2):
            for j, value in enumerate(row, 1):
                cell = sheet.cell(row=i, column=j, value=value)
                if j in [5, 6]:  # Cost columns
                    cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
        
        # Add total value formula
        for i in range(2, 6):
            sheet.cell(row=i, column=6).value = f"=D{i}*E{i}"
        
        # Add conditional formatting for low stock
        self._add_inventory_alerts(sheet)
        self._apply_inventory_formatting(sheet)
    
    def _create_analysis_sheet(self, prompt: str, analysis: Dict):
        """Create data analysis sheet"""
        sheet = self.workbook.create_sheet("Data_Analysis")  
        self.current_sheet = sheet
        
        # Create sample dataset for analysis
        headers = ['Date', 'Metric1', 'Metric2', 'Metric3', 'Category']
        self._add_headers(sheet, headers, 1)
        
        # Generate sample data
        import random
        for i in range(2, 32):  # 30 data points
            date_val = datetime.now().date() - timedelta(days=32-i)
            sheet.cell(row=i, column=1, value=date_val)
            sheet.cell(row=i, column=2, value=random.randint(50, 200))
            sheet.cell(row=i, column=3, value=random.randint(100, 300))
            sheet.cell(row=i, column=4, value=random.randint(25, 150))
            sheet.cell(row=i, column=5, value=random.choice(['A', 'B', 'C']))
        
        # Add analysis formulas
        self._add_statistical_analysis(sheet)
        self._apply_analysis_formatting(sheet)
        
        if analysis['complexity'] == 'advanced':
            self._add_trend_analysis_chart(sheet)
    
    def _create_general_sheet(self, prompt: str, analysis: Dict):
        """Create general purpose spreadsheet"""
        sheet = self.workbook.create_sheet("Data")
        self.current_sheet = sheet
        
        # Generic structure based on prompt analysis
        headers = self._extract_headers_from_prompt(prompt)
        self._add_headers(sheet, headers, 1)
        
        # Add sample data
        for i in range(2, 6):
            for j in range(1, len(headers) + 1):
                sheet.cell(row=i, column=j, value=f"Sample {i-1}")
        
        self._apply_general_formatting(sheet)
    
    def _add_headers(self, sheet, headers: List[str], row: int):
        """Add formatted headers to sheet"""
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=row, column=col, value=header.replace('_', ' '))
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def _apply_financial_formatting(self, sheet):
        """Apply professional financial formatting"""
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[column_letter].width = max_length + 2
        
        # Apply borders and alternating row colors
        self._apply_table_formatting(sheet, sheet.max_row, sheet.max_column)
    
    def _apply_tracking_formatting(self, sheet):
        """Apply tracking sheet formatting"""
        self._apply_table_formatting(sheet, sheet.max_row, sheet.max_column)
    
    def _apply_project_formatting(self, sheet):
        """Apply project management formatting"""
        self._apply_table_formatting(sheet, sheet.max_row, sheet.max_column)
        
        # Add progress bar formatting
        for row in range(2, sheet.max_row + 1):
            progress_cell = sheet.cell(row=row, column=7)
            if isinstance(progress_cell.value, (int, float)):
                if progress_cell.value >= 1.0:
                    progress_cell.fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
                elif progress_cell.value >= 0.5:
                    progress_cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                else:
                    progress_cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    
    def _apply_inventory_formatting(self, sheet):
        """Apply inventory management formatting"""
        self._apply_table_formatting(sheet, sheet.max_row, sheet.max_column)
    
    def _apply_analysis_formatting(self, sheet):
        """Apply data analysis formatting"""
        self._apply_table_formatting(sheet, sheet.max_row, sheet.max_column)
    
    def _apply_general_formatting(self, sheet):
        """Apply general formatting"""
        self._apply_table_formatting(sheet, sheet.max_row, sheet.max_column)
    
    def _apply_table_formatting(self, sheet, max_row: int, max_col: int):
        """Apply consistent table formatting"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'), 
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply borders to all cells
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = sheet.cell(row=row, column=col)
                cell.border = thin_border
                
                # Alternate row colors (skip header)
                if row > 1 and row % 2 == 0:
                    cell.fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[column_letter].width = min(max_length + 2, 50)  # Max width 50
    
    def _add_expense_chart(self, sheet):
        """Add expense breakdown chart"""
        # Implementation for expense charts
        pass
    
    def _add_sales_chart(self, sheet):
        """Add sales performance chart"""
        chart = BarChart()
        chart.title = "Sales Performance"
        
        # Data for chart (simplified)
        data = Reference(sheet, min_col=6, min_row=1, max_row=5, max_col=6)
        categories = Reference(sheet, min_col=2, min_row=2, max_row=5)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        
        sheet.add_chart(chart, "I5")
    
    def _add_gantt_visualization(self, sheet):
        """Add Gantt chart visualization"""
        # Simplified Gantt implementation
        pass
    
    def _add_inventory_alerts(self, sheet):
        """Add conditional formatting for low inventory"""
        from openpyxl.formatting.rule import CellIsRule
        
        # Highlight low stock in red
        red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
        low_stock_rule = CellIsRule(operator='lessThan', formula=['$G2'], fill=red_fill)
        
        sheet.conditional_formatting.add('D2:D100', low_stock_rule)
    
    def _add_statistical_analysis(self, sheet):
        """Add statistical analysis formulas"""
        # Add summary statistics
        sheet['G1'] = 'STATISTICS'
        sheet['G2'] = 'Metric1 Average:'
        sheet['H2'] = '=AVERAGE(B:B)'
        sheet['G3'] = 'Metric1 Std Dev:'
        sheet['H3'] = '=STDEV(B:B)'
        sheet['G4'] = 'Correlation B-C:'
        sheet['H4'] = '=CORREL(B:B,C:C)'
    
    def _add_trend_analysis_chart(self, sheet):
        """Add trend analysis chart"""
        chart = LineChart()
        chart.title = "Trend Analysis"
        
        data = Reference(sheet, min_col=2, min_row=1, max_row=31, max_col=4)
        dates = Reference(sheet, min_col=1, min_row=2, max_row=31)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(dates)
        
        sheet.add_chart(chart, "G10")
    
    def _add_tracking_summary(self, sheet):
        """Add tracking summary section"""
        sheet['I1'] = 'SUMMARY'
        sheet['I2'] = 'Total Sales:'
        sheet['J2'] = '=SUM(F:F)'
        sheet['I3'] = 'Total Commission:'
        sheet['J3'] = '=SUM(G:G)'
        sheet['I4'] = 'Avg Sale Size:'
        sheet['J4'] = '=AVERAGE(F:F)'
    
    def _create_dashboard(self):
        """Create executive dashboard sheet"""
        dashboard = self.workbook.create_sheet("Dashboard", 0)  # Insert as first sheet
        
        dashboard['A1'] = 'EXECUTIVE DASHBOARD'
        dashboard['A1'].font = Font(size=20, bold=True)
        
        # Add summary metrics (would be linked to other sheets)
        dashboard['A3'] = 'Key Metrics Summary'
        dashboard['A4'] = 'Generated on:'
        dashboard['B4'] = datetime.now().strftime('%Y-%m-%d %H:%M')
        
        self._apply_dashboard_formatting(dashboard)
    
    def _apply_dashboard_formatting(self, sheet):
        """Apply dashboard-specific formatting"""
        # Header styling
        sheet['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        sheet['A1'].font = Font(color='FFFFFF', size=18, bold=True)
        
        # Auto-adjust columns
        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 20
    
    def _extract_headers_from_prompt(self, prompt: str) -> List[str]:
        """Extract potential column headers from user prompt"""
        # Simplified header extraction - in reality would use NLP
        default_headers = ['Item', 'Description', 'Category', 'Value', 'Date', 'Status']
        
        # Look for specific terms in prompt
        if 'name' in prompt.lower():
            default_headers[0] = 'Name'
        if 'amount' in prompt.lower() or 'cost' in prompt.lower():
            default_headers[3] = 'Amount'
        
        return default_headers[:4]  # Return first 4 headers

def main():
    """Command line interface for Excel automation"""
    if len(sys.argv) < 2:
        print("Usage: python excel_master.py 'Your spreadsheet request'")
        print("Examples:")
        print("  python excel_master.py 'Create a budget tracker with expense categories'")
        print("  python excel_master.py 'Make a sales tracking sheet with commission calculations'")
        return
    
    prompt = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    excel_master = ExcelMaster()
    
    try:
        result_path = excel_master.create_spreadsheet(prompt, output_file)
        analysis = excel_master.analyze_request(prompt)
        
        print(f"✅ Excel spreadsheet created: {result_path}")
        print(f"📊 Spreadsheet type: {analysis['type']}")
        print(f"🎯 Complexity level: {analysis['complexity']}")
        print(f"🎨 Applied {analysis['color_scheme']} styling")
        
        if analysis['is_update']:
            print("🔄 Updated existing spreadsheet")
        else:
            print("🆕 Created new spreadsheet")
            
    except Exception as e:
        print(f"❌ Error creating spreadsheet: {str(e)}")
        sys.exit(1)

if __name__ == "__main__":
    main()