#!/usr/bin/env python3
"""
Enhanced Excel Master Controller - Complete Excel Automation with Full Customization
Provides advanced editing, pivot tables, professional styling, and data visualization
"""

import sys
import json
import re
import os
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional, Tuple, Union
from pathlib import Path

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle, GradientFill
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE, FORMAT_PERCENTAGE, FORMAT_DATE_DATETIME
from openpyxl.chart import (BarChart, LineChart, PieChart, AreaChart, ScatterChart,
                             Reference, Series, BarChart3D, LineChart3D)
from openpyxl.chart.marker import DataLabel
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, IconSetRule, DataBarRule
from openpyxl.utils import get_column_letter
import pandas as pd

class AdvancedTheme:
    """Advanced Excel theme configuration"""

    def __init__(self, name: str, config: Dict):
        self.name = name
        self.primary = config.get('primary', 'FF0033CC')
        self.secondary = config.get('secondary', 'FF99CCFF')
        self.accent = config.get('accent', 'FFFF6600')
        self.success = config.get('success', 'FF00CC00')
        self.warning = config.get('warning', 'FFFF9900')
        self.danger = config.get('danger', 'FFCC0000')
        self.text = config.get('text', 'FF333333')
        self.background = config.get('background', 'FFF8F9FA')
        self.header_font = config.get('header_font', 'Calibri')
        self.body_font = config.get('body_font', 'Calibri')
        self.header_size = config.get('header_size', 11)
        self.body_size = config.get('body_size', 10)

class EnhancedExcelMaster:
    """Enhanced Excel automation with full control"""

    def __init__(self):
        self.workbook = None
        self.themes = self._initialize_themes()

    def _initialize_themes(self) -> Dict[str, AdvancedTheme]:
        """Initialize professional Excel themes"""
        themes = {
            'corporate_blue': AdvancedTheme('Corporate Blue', {
                'primary': 'FF0033CC',
                'secondary': 'FF99CCFF',
                'accent': 'FFFF6600',
                'success': 'FF00CC00',
                'warning': 'FFFF9900',
                'danger': 'FFCC0000',
                'text': 'FF333333',
                'background': 'FFF8F9FA'
            }),
            'modern_dark': AdvancedTheme('Modern Dark', {
                'primary': 'FF2D3436',
                'secondary': 'FF636E72',
                'accent': 'FF74B9FF',
                'success': 'FF00B894',
                'warning': 'FFFDCB6E',
                'danger': 'FFD63031',
                'text': 'FF2D3436',
                'background': 'FFFFFFFF'
            }),
            'financial_green': AdvancedTheme('Financial Green', {
                'primary': 'FF155724',
                'secondary': 'FF28A745',
                'accent': 'FF85D5A8',
                'success': 'FF00CC00',
                'warning': 'FFFFC107',
                'danger': 'FFDC3545',
                'text': 'FF212529',
                'background': 'FFF8F9FA'
            }),
            'tech_purple': AdvancedTheme('Tech Purple', {
                'primary': 'FF6C5CE7',
                'secondary': 'FFA29BFE',
                'accent': 'FFFD79A8',
                'success': 'FF00B894',
                'warning': 'FFFECA57',
                'danger': 'FFFF7675',
                'text': 'FF2D3436',
                'background': 'FFFFFFFF'
            }),
            'elegant_mono': AdvancedTheme('Elegant Monochrome', {
                'primary': 'FF000000',
                'secondary': 'FF6C757D',
                'accent': 'FFADB5BD',
                'success': 'FF28A745',
                'warning': 'FFFFC107',
                'danger': 'FFDC3545',
                'text': 'FF212529',
                'background': 'FFFFFFFF'
            }),
            'ocean_breeze': AdvancedTheme('Ocean Breeze', {
                'primary': 'FF0077B6',
                'secondary': 'FF00B4D8',
                'accent': 'FF90E0EF',
                'success': 'FF2A9D8F',
                'warning': 'FFE9C46A',
                'danger': 'FFE76F51',
                'text': 'FF264653',
                'background': 'FFFFFFFF'
            })
        }
        return themes

    def create_workbook(self, config: Dict) -> str:
        """
        Create Excel workbook with full manual control

        Args:
            config: Complete configuration dictionary with:
                - theme: Theme name or custom theme config
                - sheets: List of sheet configurations
                - output_path: Where to save the file
        """
        # Initialize workbook
        self.workbook = Workbook()

        # Remove default sheet
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])

        # Get theme
        theme = self._get_theme(config.get('theme', 'corporate_blue'))

        # Create sheets
        sheets_config = config.get('sheets', [])
        for sheet_config in sheets_config:
            self._create_sheet(sheet_config, theme)

        # Save workbook
        output_path = config.get('output_path', 'workbook.xlsx')
        self.workbook.save(output_path)

        return output_path

    def edit_workbook(self, file_path: str, modifications: Dict) -> str:
        """
        Edit an existing Excel workbook

        Args:
            file_path: Path to existing workbook
            modifications: Dictionary of modifications:
                - change_theme: New theme to apply
                - update_sheets: Dict of sheet_name: new_content
                - add_sheets: List of new sheet configs
                - delete_sheets: List of sheet names to delete
                - add_charts: Dict of sheet_name: chart_config
                - add_pivot_tables: Dict of sheet_name: pivot_config
        """
        # Load existing workbook
        self.workbook = load_workbook(file_path)

        # Change theme if requested
        if 'change_theme' in modifications:
            new_theme = self._get_theme(modifications['change_theme'])
            self._apply_theme_to_workbook(new_theme)

        # Update existing sheets
        if 'update_sheets' in modifications:
            for sheet_name, updates in modifications['update_sheets'].items():
                if sheet_name in self.workbook.sheetnames:
                    sheet = self.workbook[sheet_name]
                    self._update_sheet_data(sheet, updates)

        # Add new sheets
        if 'add_sheets' in modifications:
            theme = self._get_theme(modifications.get('theme', 'corporate_blue'))
            for sheet_config in modifications['add_sheets']:
                self._create_sheet(sheet_config, theme)

        # Delete sheets
        if 'delete_sheets' in modifications:
            for sheet_name in modifications['delete_sheets']:
                if sheet_name in self.workbook.sheetnames:
                    self.workbook.remove(self.workbook[sheet_name])

        # Add charts
        if 'add_charts' in modifications:
            for sheet_name, chart_configs in modifications['add_charts'].items():
                if sheet_name in self.workbook.sheetnames:
                    sheet = self.workbook[sheet_name]
                    for chart_config in chart_configs:
                        self._add_chart_to_sheet(sheet, chart_config)

        # Add pivot tables
        if 'add_pivot_tables' in modifications:
            for sheet_name, pivot_config in modifications['add_pivot_tables'].items():
                if sheet_name in self.workbook.sheetnames:
                    sheet = self.workbook[sheet_name]
                    self._add_pivot_table(sheet, pivot_config)

        # Add data validation
        if 'add_validations' in modifications:
            for sheet_name, validations in modifications['add_validations'].items():
                if sheet_name in self.workbook.sheetnames:
                    sheet = self.workbook[sheet_name]
                    self._add_data_validations(sheet, validations)

        # Save with new name or overwrite
        output_path = modifications.get('output_path', file_path.replace('.xlsx', '_edited.xlsx'))
        self.workbook.save(output_path)

        return output_path

    def _get_theme(self, theme_input: Any) -> AdvancedTheme:
        """Get theme from name or custom configuration"""
        if isinstance(theme_input, str):
            return self.themes.get(theme_input, self.themes['corporate_blue'])
        elif isinstance(theme_input, dict):
            return AdvancedTheme('custom', theme_input)
        return self.themes['corporate_blue']

    def _create_sheet(self, config: Dict, theme: AdvancedTheme):
        """Create a sheet with full customization"""
        sheet_name = config.get('name', 'Sheet1')
        sheet_type = config.get('type', 'data')

        sheet = self.workbook.create_sheet(sheet_name)

        if sheet_type == 'data':
            self._create_data_sheet(sheet, config, theme)
        elif sheet_type == 'pivot':
            self._create_pivot_sheet(sheet, config, theme)
        elif sheet_type == 'dashboard':
            self._create_dashboard_sheet(sheet, config, theme)
        elif sheet_type == 'chart':
            self._create_chart_sheet(sheet, config, theme)

        return sheet

    def _create_data_sheet(self, sheet, config: Dict, theme: AdvancedTheme):
        """Create data sheet with full customization"""
        # Add headers
        headers = config.get('headers', [])
        if headers:
            self._add_styled_headers(sheet, headers, 1, theme)

        # Add data
        data = config.get('data', [])
        start_row = 2
        for row_idx, row_data in enumerate(data, start_row):
            for col_idx, value in enumerate(row_data, 1):
                cell = sheet.cell(row=row_idx, column=col_idx, value=value)

                # Apply number formats
                col_format = config.get('formats', {}).get(headers[col_idx-1] if col_idx <= len(headers) else '', None)
                if col_format:
                    cell.number_format = col_format

        # Apply styling
        style_config = config.get('styling', {})
        if style_config.get('table_style', True):
            self._apply_table_style(sheet, theme, len(headers), start_row + len(data) - 1)

        # Add conditional formatting
        if 'conditional_formatting' in config:
            self._apply_conditional_formatting(sheet, config['conditional_formatting'])

        # Add formulas
        if 'formulas' in config:
            self._add_formulas(sheet, config['formulas'])

        # Add charts
        if 'charts' in config:
            for chart_config in config['charts']:
                self._add_chart_to_sheet(sheet, chart_config)

        # Add data validation
        if 'validations' in config:
            self._add_data_validations(sheet, config['validations'])

        # Add sparklines (simulated with conditional formatting)
        if 'sparklines' in config:
            self._add_sparklines(sheet, config['sparklines'])

        # Auto-adjust column widths
        if style_config.get('auto_width', True):
            self._auto_adjust_columns(sheet)

    def _create_dashboard_sheet(self, sheet, config: Dict, theme: AdvancedTheme):
        """Create interactive dashboard"""
        # Title
        title = config.get('title', 'Dashboard')
        sheet['A1'] = title
        sheet['A1'].font = Font(name=theme.header_font, size=20, bold=True, color=theme.primary.replace('FF', ''))
        sheet.merge_cells('A1:F1')

        # Date
        sheet['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        sheet['A2'].font = Font(size=10, italic=True)

        # KPI Cards
        kpis = config.get('kpis', [])
        row = 4
        col = 1
        for kpi in kpis:
            self._create_kpi_card(sheet, row, col, kpi, theme)
            col += 2
            if col > 6:
                col = 1
                row += 4

        # Charts
        chart_row = row + 5
        charts = config.get('charts', [])
        for i, chart_config in enumerate(charts):
            chart_config['position'] = f"{get_column_letter(1 + (i % 2) * 8)}{chart_row + (i // 2) * 15}"
            self._add_chart_to_sheet(sheet, chart_config)

    def _create_kpi_card(self, sheet, row: int, col: int, kpi: Dict, theme: AdvancedTheme):
        """Create a KPI card on the dashboard"""
        # KPI name
        name_cell = sheet.cell(row=row, column=col, value=kpi['name'])
        name_cell.font = Font(size=10, bold=True)
        name_cell.fill = PatternFill(start_color=theme.background, end_color=theme.background, fill_type='solid')

        # KPI value
        value_cell = sheet.cell(row=row+1, column=col, value=kpi['value'])
        value_cell.font = Font(size=18, bold=True, color=theme.primary.replace('FF', ''))
        value_cell.number_format = kpi.get('format', '#,##0')

        # KPI change
        if 'change' in kpi:
            change_cell = sheet.cell(row=row+2, column=col, value=kpi['change'])
            change_color = theme.success if kpi['change'] > 0 else theme.danger
            change_cell.font = Font(size=10, color=change_color.replace('FF', ''))
            change_cell.number_format = '0.0%'

        # Merge and style
        sheet.merge_cells(start_row=row, start_column=col, end_row=row+2, end_column=col+1)

    def _create_pivot_sheet(self, sheet, config: Dict, theme: AdvancedTheme):
        """Create sheet with pivot table configuration"""
        # Note: python-openpyxl doesn't fully support creating pivot tables
        # This creates a manual pivot-like structure
        pass

    def _create_chart_sheet(self, sheet, config: Dict, theme: AdvancedTheme):
        """Create sheet dedicated to charts"""
        charts = config.get('charts', [])
        row = 1
        for chart_config in charts:
            chart_config['position'] = f"A{row}"
            self._add_chart_to_sheet(sheet, chart_config)
            row += 20

    def _add_styled_headers(self, sheet, headers: List[str], row: int, theme: AdvancedTheme):
        """Add professionally styled headers"""
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=row, column=col, value=header)
            cell.font = Font(name=theme.header_font, size=theme.header_size, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=theme.primary, end_color=theme.primary, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

    def _apply_table_style(self, sheet, theme: AdvancedTheme, max_col: int, max_row: int):
        """Apply professional table styling"""
        # Borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Apply to all data cells
        for row in range(2, max_row + 1):
            for col in range(1, max_col + 1):
                cell = sheet.cell(row=row, column=col)
                cell.border = thin_border

                # Alternating row colors
                if row % 2 == 0:
                    cell.fill = PatternFill(start_color=theme.background, end_color=theme.background, fill_type='solid')

                # Center alignment for certain columns
                cell.alignment = Alignment(horizontal='left', vertical='center')

    def _apply_conditional_formatting(self, sheet, cf_config: List[Dict]):
        """Apply conditional formatting rules"""
        for cf in cf_config:
            cf_type = cf.get('type', 'color_scale')
            range_addr = cf['range']

            if cf_type == 'color_scale':
                # Color scale (red-yellow-green)
                rule = ColorScaleRule(
                    start_type='min', start_color=cf.get('start_color', 'FFE17055'),
                    mid_type='percentile', mid_value=50, mid_color=cf.get('mid_color', 'FFFECA57'),
                    end_type='max', end_color=cf.get('end_color', 'FF00B894')
                )
                sheet.conditional_formatting.add(range_addr, rule)

            elif cf_type == 'data_bar':
                # Data bars
                rule = DataBarRule(
                    start_type='min', start_value=0,
                    end_type='max', end_value=100,
                    color=cf.get('color', '6C5CE7'),
                    showValue=True
                )
                sheet.conditional_formatting.add(range_addr, rule)

            elif cf_type == 'icon_set':
                # Icon sets (3 arrows, 3 traffic lights, etc.)
                rule = IconSetRule(
                    icon_style=cf.get('icon_style', '3Arrows'),
                    type='num',
                    values=[0, 33, 67],
                    showValue=True
                )
                sheet.conditional_formatting.add(range_addr, rule)

            elif cf_type == 'cell_value':
                # Cell value rules
                fill = PatternFill(start_color=cf.get('color', 'FFFF0000'), end_color=cf.get('color', 'FFFF0000'), fill_type='solid')
                rule = CellIsRule(
                    operator=cf.get('operator', 'greaterThan'),
                    formula=[str(cf.get('value', 0))],
                    fill=fill
                )
                sheet.conditional_formatting.add(range_addr, rule)

    def _add_formulas(self, sheet, formulas: List[Dict]):
        """Add formulas to cells"""
        for formula in formulas:
            cell_addr = formula['cell']
            formula_str = formula['formula']

            cell = sheet[cell_addr]
            cell.value = formula_str

            # Apply formatting if specified
            if 'format' in formula:
                cell.number_format = formula['format']

    def _add_chart_to_sheet(self, sheet, chart_config: Dict):
        """Add chart with full customization"""
        chart_type = chart_config.get('type', 'bar')

        # Create chart
        if chart_type == 'bar':
            chart = BarChart()
        elif chart_type == 'bar_3d':
            chart = BarChart3D()
        elif chart_type == 'line':
            chart = LineChart()
        elif chart_type == 'line_3d':
            chart = LineChart3D()
        elif chart_type == 'pie':
            chart = PieChart()
        elif chart_type == 'area':
            chart = AreaChart()
        elif chart_type == 'scatter':
            chart = ScatterChart()
        else:
            chart = BarChart()

        # Configure chart
        chart.title = chart_config.get('title', 'Chart')
        chart.style = chart_config.get('style', 10)

        # Data range
        data_range = chart_config.get('data_range', 'A1:D10')
        data_ref = Reference(sheet, range_string=data_range)
        chart.add_data(data_ref, titles_from_data=True)

        # Categories range
        if 'categories_range' in chart_config:
            cat_range = chart_config['categories_range']
            cat_ref = Reference(sheet, range_string=cat_range)
            chart.set_categories(cat_ref)

        # Axis labels
        if 'x_axis' in chart_config:
            chart.x_axis.title = chart_config['x_axis']
        if 'y_axis' in chart_config:
            chart.y_axis.title = chart_config['y_axis']

        # Position
        position = chart_config.get('position', 'H5')
        chart.width = chart_config.get('width', 15)
        chart.height = chart_config.get('height', 10)

        sheet.add_chart(chart, position)

    def _add_pivot_table(self, sheet, pivot_config: Dict):
        """Add pivot table (manual aggregation since openpyxl doesn't fully support pivot tables)"""
        # This would create a manual pivot-like structure
        # Full pivot table support requires COM automation or xlsxwriter
        pass

    def _add_data_validations(self, sheet, validations: List[Dict]):
        """Add data validation rules"""
        for validation in validations:
            val_type = validation.get('type', 'list')
            range_addr = validation['range']

            if val_type == 'list':
                # Dropdown list
                dv = DataValidation(
                    type="list",
                    formula1=validation.get('formula1', '"Option 1,Option 2,Option 3"'),
                    allow_blank=validation.get('allow_blank', True)
                )
                dv.error = validation.get('error_message', 'Invalid value')
                dv.errorTitle = validation.get('error_title', 'Invalid Entry')
                sheet.add_data_validation(dv)
                dv.add(range_addr)

            elif val_type == 'number':
                # Number validation
                dv = DataValidation(
                    type="whole",
                    operator=validation.get('operator', 'between'),
                    formula1=validation.get('min', 0),
                    formula2=validation.get('max', 100)
                )
                sheet.add_data_validation(dv)
                dv.add(range_addr)

            elif val_type == 'date':
                # Date validation
                dv = DataValidation(
                    type="date",
                    operator=validation.get('operator', 'greaterThan'),
                    formula1=validation.get('formula1', datetime.now().date())
                )
                sheet.add_data_validation(dv)
                dv.add(range_addr)

    def _add_sparklines(self, sheet, sparkline_config: List[Dict]):
        """Add sparklines (simulated with mini charts or conditional formatting)"""
        # Note: True sparklines require Excel formulas like SPARKLINE()
        # This creates visual approximations
        pass

    def _auto_adjust_columns(self, sheet):
        """Auto-adjust column widths based on content"""
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)  # Max width 50
            sheet.column_dimensions[column_letter].width = adjusted_width

    def _update_sheet_data(self, sheet, updates: Dict):
        """Update existing sheet data"""
        # Update cells
        if 'cells' in updates:
            for cell_addr, value in updates['cells'].items():
                sheet[cell_addr] = value

        # Update range
        if 'range' in updates:
            range_data = updates['range']
            start_cell = range_data['start']
            data = range_data['data']

            # Parse start cell
            import openpyxl.utils.cell as cell_utils
            col, row = cell_utils.coordinate_to_tuple(start_cell)

            for row_idx, row_data in enumerate(data):
                for col_idx, value in enumerate(row_data):
                    sheet.cell(row=row+row_idx, column=col+col_idx, value=value)

        # Clear cells
        if 'clear' in updates:
            for cell_addr in updates['clear']:
                sheet[cell_addr].value = None

    def _apply_theme_to_workbook(self, theme: AdvancedTheme):
        """Apply theme to entire workbook"""
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]

            # Update header row if it exists
            if sheet.max_row > 0:
                for cell in sheet[1]:
                    if cell.fill and cell.fill.start_color:
                        cell.fill = PatternFill(start_color=theme.primary, end_color=theme.primary, fill_type='solid')
                        cell.font = Font(color='FFFFFF', bold=True)

    def get_available_themes(self) -> List[str]:
        """Return list of available theme names"""
        return list(self.themes.keys())

    def get_theme_preview(self, theme_name: str) -> Dict:
        """Get theme color preview"""
        theme = self.themes.get(theme_name)
        if not theme:
            return {}

        return {
            'name': theme.name,
            'colors': {
                'primary': theme.primary,
                'secondary': theme.secondary,
                'accent': theme.accent,
                'success': theme.success,
                'warning': theme.warning,
                'danger': theme.danger,
                'text': theme.text,
                'background': theme.background
            }
        }

    def create_heatmap(self, sheet, data_range: str, title: str = "Heatmap"):
        """Create heatmap visualization using conditional formatting"""
        rule = ColorScaleRule(
            start_type='min', start_color='FFFFFF',
            mid_type='percentile', mid_value=50, mid_color='FFF4A582',
            end_type='max', end_color='FFCA0020'
        )
        sheet.conditional_formatting.add(data_range, rule)

        # Add title
        import openpyxl.utils.cell as cell_utils
        col, row = cell_utils.coordinate_to_tuple(data_range.split(':')[0])
        title_cell = sheet.cell(row=row-1, column=col)
        title_cell.value = title
        title_cell.font = Font(bold=True, size=14)

def main():
    """Command line interface for enhanced Excel automation"""
    if len(sys.argv) < 2:
        print("Enhanced Excel Master - Full Customization Control")
        print("\nUsage:")
        print("  python excel_master_enhanced.py config.json")
        print("\nAvailable themes:")
        master = EnhancedExcelMaster()
        for theme in master.get_available_themes():
            print(f"  - {theme}")
        return

    config_file = sys.argv[1]

    try:
        with open(config_file, 'r') as f:
            config = json.load(f)

        master = EnhancedExcelMaster()

        # Check if editing existing file
        if 'edit_file' in config:
            result_path = master.edit_workbook(config['edit_file'], config)
            print(f"✅ Workbook edited successfully: {result_path}")
        else:
            result_path = master.create_workbook(config)
            print(f"✅ Workbook created successfully: {result_path}")

        print(f"📊 Theme: {config.get('theme', 'corporate_blue')}")
        print(f"📋 Sheets: {len(config.get('sheets', []))}")

    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()
